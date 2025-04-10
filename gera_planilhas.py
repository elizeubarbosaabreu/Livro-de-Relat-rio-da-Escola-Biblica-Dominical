import os
import calendar
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Configurações gerais
print("Não digite planilhas existentes, senão serão substituídas.")
inicio = int(input("Digite o ano inicial: "))
fim = int(input("Digite o ano final: "))
ANOS = range(inicio, fim+1)

def carregar_classes(caminho_arquivo="classes.txt"):
    try:
        with open(caminho_arquivo, encoding="utf-8") as f:
            return [linha.strip() for linha in f if linha.strip()]
    except FileNotFoundError:
        messagebox.showerror("Erro", f"Arquivo '{caminho_arquivo}' não encontrado.")
        return []

CLASSES = carregar_classes()


COLUNAS = ["Classes", "Matriculados", "Ausentes", "Presentes", "Visitantes", "Total", "Bíblias", "Revistas", "Ofertas", "% de Presença"]
MESES_PT = {
    1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"
}

# Estilos
header_font = Font(bold=True)
header_fill = PatternFill("solid", fgColor="C0C0C0")
center_align = Alignment(horizontal="center")
fill_zebra = PatternFill("solid", fgColor="F2F2F2")
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def domingos_do_mes(ano, mes):
    dias = calendar.monthrange(ano, mes)[1]
    return [day for day in range(1, dias+1) if datetime(ano, mes, day).weekday() == 6]

def cria_planilha_para_domingo(caminho, data):
    nome_arquivo = data.strftime('%d') + f"_{MESES_PT[data.month]}_{data.year}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Presença"

    # Título mesclado
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Relatório da Escola Bíblica Dominical em {data.strftime('%d/%m/%Y')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # Cabeçalhos
    for i, coluna in enumerate(COLUNAS, 1):
        celula = ws.cell(row=2, column=i, value=coluna)
        celula.font = header_font
        celula.fill = header_fill
        # Alinhar verticalmente para certas colunas
        vertical_cols = ["Matriculados", "Ausentes", "Presentes", "Visitantes", "Total", "Bíblias", "Revistas", "Ofertas", "% de Presença"]
        if coluna in vertical_cols:
            celula.alignment = Alignment(horizontal="center", vertical="center", textRotation=90)
        else:
            celula.alignment = center_align

        celula.border = border

    # Preenchimento das classes
    for i, classe in enumerate(CLASSES, start=3):
        ws.cell(row=i, column=1, value=classe)
        ws[f"D{i}"] = f"=B{i}-C{i}"  # Presentes
        ws[f"F{i}"] = f"=D{i}+E{i}"  # Total
        ws[f"J{i}"] = f"=IF(B{i}>0,D{i}/B{i},0)"
        ws[f"I{i}"].number_format = 'R$ #,##0.00' # Valor monetário
        ws[f"J{i}"].number_format = '0.00%'
        
        # Zebra striping
        if (i - 3) % 2 == 1:
            for j in range(1, 11):
                ws.cell(row=i, column=j).fill = fill_zebra

        # Aplicar bordas
        for j in range(1, 11):
            ws.cell(row=i, column=j).border = border

    # Linha de total geral
    linha_total = len(CLASSES) + 3
    ws.cell(row=linha_total, column=1, value="Total Geral")
    for col in range(2, 10):
        letra = get_column_letter(col)
        ws.cell(row=linha_total, column=col, value=f"=SUM({letra}3:{letra}{linha_total-1})")
        ws.cell(row=linha_total, column=col).font = Font(bold=True)
        ws.cell(row=linha_total, column=col).alignment = center_align
        ws.cell(row=linha_total, column=col).border = border

    # % de presença total geral
    ws[f"J{linha_total}"] = f"=IF(B{linha_total}>0,D{linha_total}/B{linha_total},0)"
    ws[f"J{linha_total}"].number_format = '0.00%'
    ws[f"J{linha_total}"].font = Font(bold=True)
    ws[f"J{linha_total}"].alignment = center_align
    ws[f"J{linha_total}"].border = border

    # Formatação condicional nas classes
    ws[f"I{linha_total}"].number_format = 'R$ #,##0.00' # Valor monetário
    ws.conditional_formatting.add(
        f"J3:J{linha_total-1}",
        CellIsRule(operator='greaterThanOrEqual', formula=['0.9'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        f"J3:J{linha_total-1}",
        CellIsRule(operator='lessThan', formula=['0.7'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
    )

    # Ajuste automático de largura das colunas
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            len(str(ws.cell(row=row, column=col_idx).value)) if ws.cell(row=row, column=col_idx).value else 0
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(os.path.join(caminho, nome_arquivo))
    

def criar_estrutura_ebd(base_dir="Relatorios_EBD"):
    os.makedirs(base_dir, exist_ok=True)
    for ano in ANOS:
        ano_dir = os.path.join(base_dir, str(ano))
        os.makedirs(ano_dir, exist_ok=True)
        for mes in range(1, 13):
            nome_mes = MESES_PT[mes]
            mes_dir = os.path.join(ano_dir, nome_mes)
            os.makedirs(mes_dir, exist_ok=True)
            for dia in domingos_do_mes(ano, mes):
                data_domingo = datetime(ano, mes, dia)
                cria_planilha_para_domingo(mes_dir, data_domingo)

# Executar tudo
criar_estrutura_ebd()
print("Planilhas Criadas com sucesso")

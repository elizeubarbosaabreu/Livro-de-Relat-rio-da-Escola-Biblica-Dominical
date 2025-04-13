from fpdf import FPDF
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

# === CONFIGURAÇÕES ===
DIRETORIO = "Relatorios_EBD"
CLASSES_TXT = "classes.txt"
IGREJA_TXT = "igreja.txt"

MESES_PT = {
    1: "janeiro", 2: "fevereiro", 3: "março", 4: "abril",
    5: "maio", 6: "junho", 7: "julho", 8: "agosto",
    9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"
}

CAMPOS = [
    "Classe", "Matriculados", "Ausentes", "Presentes", "Visitantes",
    "Total", "Bíblias", "Revistas", "Ofertas", "% de Presença"
]

# === FUNÇÕES AUXILIARES ===
def ler_igreja():
    try:
        with open(IGREJA_TXT, "r", encoding="utf-8") as f:
            return f.read().strip()
    except FileNotFoundError:
        return "[Nome da Igreja]"

def ler_classes():
    try:
        with open(CLASSES_TXT, "r", encoding="utf-8") as f:
            return [linha.strip() for linha in f if linha.strip()]
    except FileNotFoundError:
        return []

def carregar_dados(data):
    mes = MESES_PT[data.month].lower()
    ano = data.year
    dia = f"{data.day:02d}"
    caminho = os.path.join(DIRETORIO, str(ano), mes, f"{dia}_{mes}_{ano}.xlsx")
    if not os.path.exists(caminho):
        return {}

    wb = load_workbook(caminho, data_only=True)
    ws = wb.active
    dados = {}

    for linha in ws.iter_rows(min_row=3, values_only=True):
        if not linha[0]:
            continue
        dados[linha[0]] = {
            "Matriculados": linha[1],
            "Ausentes": linha[2],
            "Presentes": linha[3],
            "Visitantes": linha[4],
            "Total": linha[5],
            "Bíblias": linha[6],
            "Revistas": linha[7],
            "Ofertas": linha[8],
            "% de Presença": round(float(linha[9]) * 100) if linha[9] is not None else 0,
        }
    return dados

# === FUNÇÃO PRINCIPAL DE GERAÇÃO ===
def gerar_pdf(DATA_ATUAL, DATA_ANTERIOR):
    igreja = ler_igreja()
    classes = ler_classes()
    dados_atuais = carregar_dados(DATA_ATUAL)

    if not dados_atuais:
        print(f"❌ Nenhum dado encontrado para {DATA_ATUAL.strftime('%d/%m/%Y')}. Relatório não gerado.")
        return

    dados_anteriores = carregar_dados(DATA_ANTERIOR)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(0, 10, "Relatório da Escola Bíblica Dominical", ln=True, align='C')
    pdf.set_font("Arial", '', 12)
    pdf.cell(0, 10, f"{igreja} aos {DATA_ATUAL.day} de {MESES_PT[DATA_ATUAL.month]} de {DATA_ATUAL.year}", ln=True, align='C')
    pdf.ln(10)

    total_geral_atual = {campo: 0 for campo in CAMPOS[1:]}
    total_geral_anterior = {campo: 0 for campo in CAMPOS[1:]}

    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 10, "Resumo por Classe:", ln=True)
    pdf.set_font("Arial", '', 10)

    for classe in classes:
        atual = dados_atuais.get(classe, {})
        anterior = dados_anteriores.get(classe, {})

        perc = atual.get("% de Presença", 0)

        # Define a cor de fundo com base no percentual de presença
        if perc >= 80:
            pdf.set_fill_color(144, 238, 144)  # verde
        elif perc < 60:
            pdf.set_fill_color(255, 99, 71)    # vermelho
        else:
            pdf.set_fill_color(255, 255, 153)  # amarelo
        fill = True

        linha = f"{classe}: Mat {atual.get('Matriculados', 0)}, Aus {atual.get('Ausentes', 0)}, " \
                f"Pres {atual.get('Presentes', 0)}, Vis {atual.get('Visitantes', 0)}, Total {atual.get('Total', 0)}, " \
                f"Bíblias {atual.get('Bíblias', 0)}, Revistas {atual.get('Revistas', 0)}, " \
                f"Presença {perc}%"
        pdf.cell(0, 8, linha, ln=True, fill=fill)

        for campo in CAMPOS[1:]:
            try:
                total_geral_atual[campo] += float(atual.get(campo, 0))
                total_geral_anterior[campo] += float(anterior.get(campo, 0))
            except:
                pass

    pdf.ln(8)
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 10, "Comparativo dos Totais Gerais:", ln=True)
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(65, 8, f"Domingo atual: {DATA_ATUAL.strftime('%d/%m/%Y')}", border=1)
    pdf.cell(65, 8, f"Domingo anterior: {DATA_ANTERIOR.strftime('%d/%m/%Y')}", border=1, ln=True)

    pdf.set_font("Arial", '', 10)
    for campo in CAMPOS[1:]:
        atual = total_geral_atual.get(campo, 0)
        anterior = total_geral_anterior.get(campo, 0)

        if campo == "Ofertas":
            atual_str = f"R$ {atual:.2f}"
            anterior_str = f"R$ {anterior:.2f}"
        elif campo == "% de Presença":
            atual = round(total_geral_atual["Presentes"] / total_geral_atual["Matriculados"] * 100) if total_geral_atual["Matriculados"] else 0
            anterior = round(total_geral_anterior["Presentes"] / total_geral_anterior["Matriculados"] * 100) if total_geral_anterior["Matriculados"] else 0
            atual_str = f"{atual}%"
            anterior_str = f"{anterior}%"
        else:
            atual_str = str(int(atual))
            anterior_str = str(int(anterior))

        pdf.cell(65, 8, f"{campo}: {atual_str}", border=1)
        pdf.cell(65, 8, f"{campo}: {anterior_str}", border=1, ln=True)

    pdf.ln(15)
    pdf.cell(0, 8, "Secretário: ____________________________         Superintendente: __________________________", ln=True)

    mes_nome = MESES_PT[DATA_ATUAL.month].lower()
    ano = DATA_ATUAL.year
    caminho_pdf = os.path.join(DIRETORIO, str(ano), mes_nome)
    os.makedirs(caminho_pdf, exist_ok=True)

    nome_arquivo = f"relatorio_ebd_{DATA_ATUAL.strftime('%d_%m_%Y')}.pdf"
    caminho_completo = os.path.join(caminho_pdf, nome_arquivo)
    pdf.output(caminho_completo)
    print(f"✅ Relatório gerado: {caminho_completo}")

# === EXECUÇÃO PRINCIPAL ===
if __name__ == "__main__":
    ano = int(input("Digite o ano desejado: "))
    data = datetime(ano, 1, 1)

    # Avança até o primeiro domingo do ano
    while data.weekday() != 6:
        data += timedelta(days=1)

    hoje = datetime.today()

    while data.year == ano:
        DATA_ATUAL = datetime(data.year, data.month, data.day)
        DATA_ANTERIOR = DATA_ATUAL - timedelta(days=7)

        gerar_pdf(DATA_ATUAL, DATA_ANTERIOR)

        # Se atingiu o domingo atual (no ano atual), parar
        if ano == hoje.year and data.date() >= hoje.date():
            break

        data += timedelta(days=7)

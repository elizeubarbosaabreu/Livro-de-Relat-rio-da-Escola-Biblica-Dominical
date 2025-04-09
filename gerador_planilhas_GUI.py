import os
import calendar
from tkinter import Tk, Label, Button, ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_DIR = "Relatorios_EBD"
CLASSES = [
    "Diretoria", "Abraão", "Samuel", "Ana", "Débora", "Venc. em Cristo",
    "Miriã", "Lírio dos Vales", "Gideões", "Cam. p/ o Céu",
    "Rosa de Saron", "Sold. de Cristo", "Querubins"
]

CAMPOS = [
    "Classe", "Matriculados", "Ausentes", "Presentes", "Visitantes",
    "Total", "Bíblias", "Revistas", "Ofertas", "% de Presença"
]

# Tradução dos meses
MESES_PT = {
    "janeiro": "January", "fevereiro": "February", "março": "March",
    "abril": "April", "maio": "May", "junho": "June",
    "julho": "July", "agosto": "August", "setembro": "September",
    "outubro": "October", "novembro": "November", "dezembro": "December"
}

MESES_LISTA_PT = list(MESES_PT.keys())

bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center")
fill_header = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

def criar_planilha(mes_pt: str, ano: int):
    mes_en = MESES_PT[mes_pt.lower()]
    mes_dir = os.path.join(BASE_DIR, str(ano), mes_pt.lower())
    os.makedirs(mes_dir, exist_ok=True)

    num_mes = list(calendar.month_name).index(mes_en)

    dias_domingo = [
        dia for dia in range(1, calendar.monthrange(ano, num_mes)[1] + 1)
        if calendar.weekday(ano, num_mes, dia) == 6
    ]

    for dia in dias_domingo:
        nome_arquivo = f"{dia:02d}_{mes_pt}_{ano}.xlsx"
        caminho = os.path.join(mes_dir, nome_arquivo)
        wb = Workbook()
        ws = wb.active
        ws.title = "Presença"

        for col, campo in enumerate(CAMPOS, 1):
            cel = ws.cell(row=2, column=col, value=campo)
            cel.font = bold_font
            cel.alignment = center_align
            cel.fill = fill_header

        for i, classe in enumerate(CLASSES, start=3):
            ws.cell(row=i, column=1, value=classe)
            for j in range(2, len(CAMPOS) + 1):
                cel = ws.cell(row=i, column=j)
                cel.alignment = center_align

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(12, max_length + 2)

        wb.save(caminho)

def gerar_planilhas():
    try:
        ano = int(ano_var.get())
        mes_pt = mes_var.get()
        criar_planilha(mes_pt, ano)
        messagebox.showinfo("Sucesso", f"Planilhas de {mes_pt.capitalize()} de {ano} geradas com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao gerar planilhas: {str(e)}")

root = Tk()
root.title("Gerador de Planilhas da EBD")
root.geometry("360x200")
root.configure(bg="#f0f0f0")

Label(root, text="Selecione o Ano:", font=("Arial", 11), bg="#f0f0f0").pack(pady=5)
ano_var = ttk.Combobox(root, values=[str(a) for a in range(2020, 2031)], state="readonly")
ano_var.set("2025")
ano_var.pack()

Label(root, text="Selecione o Mês:", font=("Arial", 11), bg="#f0f0f0").pack(pady=5)
mes_var = ttk.Combobox(root, values=[m.capitalize() for m in MESES_LISTA_PT], state="readonly")
mes_var.set("Abril")
mes_var.pack()

Button(root, text="Gerar Planilhas", command=gerar_planilhas, width=25, bg="#4CAF50", fg="white").pack(pady=15)

root.mainloop()

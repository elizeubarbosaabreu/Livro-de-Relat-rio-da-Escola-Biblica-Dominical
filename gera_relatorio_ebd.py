import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from math import isclose

BASE_DIR = "Relatorios_EBD"
CLASSES = [
    "Diretoria", "Abraão", "Samuel", "Ana", "Débora", "Venc. em Cristo",
    "Miriã", "Lírio dos Vales", "Gideões", "Cam. p/ o Céu",
    "Rosa de Saron", "Sold. de Cristo", "Querubins"
]
COLUNAS_RELEVANTES = {
    "Matriculados": 2, "Ausentes": 3, "Presentes": 4,
    "Visitantes": 5, "Total": 6, "Bíblias": 7,
    "Revistas": 8, "Ofertas": 9, "% de Presença": 10
}
TRIMESTRES = {
    "1º Trimestre": ["janeiro", "fevereiro", "março"],
    "2º Trimestre": ["abril", "maio", "junho"],
    "3º Trimestre": ["julho", "agosto", "setembro"],
    "4º Trimestre": ["outubro", "novembro", "dezembro"]
}

bold_center = Font(bold=True)
center_align = Alignment(horizontal="center")
rotated_align = Alignment(horizontal="center", vertical="center", textRotation=90)
currency_format = 'R$ #,##0.00'
percent_format = '0.00%'

def inicializa_linha_totais():
    return {col: 0 for col in COLUNAS_RELEVANTES}

def somar_linha(origem, destino):
    for k in COLUNAS_RELEVANTES:
        destino[k] += origem[k]

def obter_dados_por_classe(arquivo):
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active
    dados = {classe: inicializa_linha_totais() for classe in CLASSES}
    for row in ws.iter_rows(min_row=3, max_row=3+len(CLASSES)-1):
        nome_classe = row[0].value
        if nome_classe not in dados:
            continue
        for k, idx in COLUNAS_RELEVANTES.items():
            celula = row[idx-1].value
            valor = float(celula) if celula not in (None, '') else 0
            dados[nome_classe][k] += valor
    return dados

def dividir_e_arredondar(valor, divisor):
    if divisor == 0:
        return 0
    if isinstance(valor, float) and isclose(valor % 1, 0.5, abs_tol=1e-6):
        return int(round(valor))
    return int(round(valor / divisor))

def criar_relatorio(nome, dados, caminho_destino, contagem_domingos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    titulo = f"Relatório do {nome.replace('.xlsx', '')} de {os.path.basename(caminho_destino)} da EBD"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS_RELEVANTES)+1)
    celula_titulo = ws.cell(row=1, column=1, value=titulo)
    celula_titulo.font = Font(bold=True, size=14)
    celula_titulo.alignment = Alignment(horizontal="center")
    cabecalho = ["Classe"] + list(COLUNAS_RELEVANTES.keys())
    for col, texto in enumerate(cabecalho, 1):
        cel = ws.cell(row=2, column=col, value=texto)
        cel.font = bold_center
        cel.alignment = rotated_align if texto != "Classe" else center_align
    for i, classe in enumerate(CLASSES, start=3):
        ws.cell(row=i, column=1, value=classe)
        for j, k in enumerate(COLUNAS_RELEVANTES.keys(), start=2):
            valor = dados[classe][k]
            cel = ws.cell(row=i, column=j)
            if k == "Ofertas":
                cel.value = valor
                cel.number_format = currency_format
            elif "%" in k:
                media = valor / contagem_domingos if contagem_domingos > 0 else 0
                cel.value = media
                cel.number_format = percent_format
            else:
                cel.value = dividir_e_arredondar(valor, contagem_domingos)
            cel.alignment = center_align
    linha_total = len(CLASSES) + 3
    ws.cell(row=linha_total, column=1, value="Total Geral").font = bold_center
    for j, k in enumerate(COLUNAS_RELEVANTES.keys(), start=2):
        total = sum(dados[classe][k] for classe in CLASSES)
        cel = ws.cell(row=linha_total, column=j)
        if k == "Ofertas":
            cel.value = total
            cel.number_format = currency_format
        elif k == "% de Presença":
            soma_matriculados = sum(dados[classe]["Matriculados"] for classe in CLASSES)
            soma_presentes = sum(dados[classe]["Presentes"] for classe in CLASSES)
            porcentagem = soma_presentes / soma_matriculados if soma_matriculados > 0 else 0
            cel.value = porcentagem
            cel.number_format = percent_format
        else:
            cel.value = dividir_e_arredondar(total, contagem_domingos)
        cel.font = bold_center
        cel.alignment = center_align
    faixa = f"J3:J{linha_total - 1}"
    ws.conditional_formatting.add(faixa,
        CellIsRule(operator='greaterThanOrEqual', formula=['0.8'],
                   fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
    ws.conditional_formatting.add(faixa,
        CellIsRule(operator='lessThan', formula=['0.6'],
                   fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 25)
    wb.save(os.path.join(caminho_destino, nome))

def gerar_relatorio_mensal(ano, mes):
    caminho_mes = os.path.join(BASE_DIR, str(ano), mes)
    if not os.path.exists(caminho_mes):
        print(f"❌ Mês '{mes}' não encontrado.")
        return
    dados_mes = {classe: inicializa_linha_totais() for classe in CLASSES}
    contagem = 0
    for arquivo in os.listdir(caminho_mes):
        if arquivo.endswith(".xlsx"):
            caminho = os.path.join(caminho_mes, arquivo)
            dados = obter_dados_por_classe(caminho)
            for classe in CLASSES:
                somar_linha(dados[classe], dados_mes[classe])
            contagem += 1
    if contagem > 0:
        criar_relatorio(f"Relatório de {mes} de {ano}.xlsx", dados_mes, caminho_mes, contagem)
        print(f"✅ Relatório mensal de {mes.capitalize()} gerado com sucesso.")
    else:
        print(f"❌ Nenhuma planilha encontrada para {mes}.")

def gerar_relatorios_anuais_e_trimestrais(ano_escolhido):
    dados_ano = {classe: inicializa_linha_totais() for classe in CLASSES}
    planilhas_ano = 0
    ano_dir = os.path.join(BASE_DIR, str(ano_escolhido))
    for trimestre, meses in TRIMESTRES.items():
        dados_trimestre = {classe: inicializa_linha_totais() for classe in CLASSES}
        contagem_trimestre = 0
        for mes in meses:
            mes_dir = os.path.join(ano_dir, mes)
            if not os.path.exists(mes_dir):
                continue
            for arquivo in os.listdir(mes_dir):
                if arquivo.endswith(".xlsx"):
                    caminho = os.path.join(mes_dir, arquivo)
                    dados = obter_dados_por_classe(caminho)
                    for classe in CLASSES:
                        somar_linha(dados[classe], dados_trimestre[classe])
                        somar_linha(dados[classe], dados_ano[classe])
                    contagem_trimestre += 1
                    planilhas_ano += 1
        criar_relatorio(f"{trimestre}.xlsx", dados_trimestre, ano_dir, contagem_trimestre)
    criar_relatorio("Relatório_Anual.xlsx", dados_ano, ano_dir, planilhas_ano)
    print(f"✅ Relatórios trimestrais e anual gerados com sucesso para {ano_escolhido}.")

# Menu interativo
def menu():
    try:
        print("\n=== GERADOR DE RELATÓRIOS EBD ===")
        ano = int(input("Digite o ano (ex: 2024): "))
        print("\nEscolha o tipo de relatório:")
        print("1 - Relatório Mensal")
        print("2 - Relatórios Trimestrais e Anual")
        print("3 - Ambos (Mensal + Trimestre + Anual)")
        opcao = input("Opção: ")

        if opcao == "1" or opcao == "3":
            mes = input("Digite o nome do mês (ex: janeiro): ").strip().lower()
            gerar_relatorio_mensal(ano, mes)

        if opcao == "2" or opcao == "3":
            gerar_relatorios_anuais_e_trimestrais(ano)

    except ValueError:
        print("❌ Entrada inválida. Use números inteiros para o ano.")
    except Exception as e:
        print(f"❌ Erro inesperado: {e}")

# Execução
if __name__ == "__main__":
    menu()

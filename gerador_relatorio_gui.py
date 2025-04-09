import os
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from math import isclose

BASE_DIR = "Relatorios_EBD"
CLASSES = [
    "Diretoria", "Abraão", "Samuel", "Ana", "Débora", "Venc. em Cristo",
    "Miriã", "Lírio dos Vales", "Gideões", "Cam. p/ o Céu",
    "Rosa de Saron", "Sold. de Cristo", "Querubins"
]
COLUNAS_RELEVANTES = {
    "Matriculados": 2, "Ausentes": 3, "Presentes": 4,
    "Visitantes": 5, "Total": 6, "Bíblias": 7,
    "Revistas": 8, "Ofertas": 9, "% de Presença": 10
}
TRIMESTRES = {
    "1º Trimestre": ["janeiro", "fevereiro", "março"],
    "2º Trimestre": ["abril", "maio", "junho"],
    "3º Trimestre": ["julho", "agosto", "setembro"],
    "4º Trimestre": ["outubro", "novembro", "dezembro"]
}

bold_center = Font(bold=True)
center_align = Alignment(horizontal="center")
rotated_align = Alignment(horizontal="center", vertical="center", textRotation=90)
currency_format = 'R$ #,##0.00'
percent_format = '0.00%'

MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]


def inicializa_linha():
    return {col: 0 for col in COLUNAS_RELEVANTES}

def somar_linha(origem, destino):
    for k in COLUNAS_RELEVANTES:
        destino[k] += origem[k]

def obter_dados_por_classe(arquivo):
    wb = load_workbook(arquivo, data_only=True)
    ws = wb.active
    dados = {classe: inicializa_linha() for classe in CLASSES}
    for row in ws.iter_rows(min_row=3, max_row=3+len(CLASSES)-1):
        nome = row[0].value
        if nome not in dados:
            continue
        for k, idx in COLUNAS_RELEVANTES.items():
            valor = row[idx-1].value
            valor = float(valor) if valor not in (None, '') else 0
            dados[nome][k] += valor
    return dados

def dividir(valor, divisor):
    if divisor == 0:
        return 0
    if isinstance(valor, float) and isclose(valor % 1, 0.5, abs_tol=1e-6):
        return int(round(valor))
    return int(round(valor / divisor))

def criar_relatorio(nome, dados, destino, domingos):
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    titulo = f"Relatório do {nome.replace('.xlsx', '')} de {os.path.basename(destino)} da EBD"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUNAS_RELEVANTES)+1)
    cel_titulo = ws.cell(row=1, column=1, value=titulo)
    cel_titulo.font = Font(bold=True, size=14)
    cel_titulo.alignment = Alignment(horizontal="center")

    cabecalho = ["Classe"] + list(COLUNAS_RELEVANTES.keys())
    for col, texto in enumerate(cabecalho, 1):
        cel = ws.cell(row=2, column=col, value=texto)
        cel.font = bold_center
        cel.alignment = rotated_align if texto != "Classe" else center_align

    for i, classe in enumerate(CLASSES, start=3):
        ws.cell(row=i, column=1, value=classe)
        for j, k in enumerate(COLUNAS_RELEVANTES.keys(), start=2):
            valor = dados[classe][k]
            cel = ws.cell(row=i, column=j)
            if k == "Ofertas":
                cel.value = valor
                cel.number_format = currency_format
            elif "%" in k:
                cel.value = valor / domingos if domingos > 0 else 0
                cel.number_format = percent_format
            else:
                cel.value = dividir(valor, domingos)
            cel.alignment = center_align

    linha_total = len(CLASSES) + 3
    ws.cell(row=linha_total, column=1, value="Total Geral").font = bold_center

    for j, k in enumerate(COLUNAS_RELEVANTES.keys(), start=2):
        total = sum(dados[c][k] for c in CLASSES)
        cel = ws.cell(row=linha_total, column=j)
        if k == "Ofertas":
            cel.value = total
            cel.number_format = currency_format
        elif k == "% de Presença":
            mat = sum(dados[c]["Matriculados"] for c in CLASSES)
            pre = sum(dados[c]["Presentes"] for c in CLASSES)
            cel.value = pre / mat if mat > 0 else 0
            cel.number_format = percent_format
        else:
            cel.value = dividir(total, domingos)
        cel.font = bold_center
        cel.alignment = center_align

    faixa = f"J3:J{linha_total - 1}"
    ws.conditional_formatting.add(faixa, CellIsRule(operator='greaterThanOrEqual', formula=['0.8'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
    ws.conditional_formatting.add(faixa, CellIsRule(operator='lessThan', formula=['0.6'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 25)

    wb.save(os.path.join(destino, nome))

def gerar_mensal():
    ano = entry_ano.get()
    mes = combo_mes.get()
    dir_mes = os.path.join(BASE_DIR, ano, mes)
    if not os.path.exists(dir_mes):
        messagebox.showerror("Erro", f"Diretório {dir_mes} não encontrado.")
        return

    dados = {c: inicializa_linha() for c in CLASSES}
    contagem = 0
    for arq in os.listdir(dir_mes):
        if arq.endswith(".xlsx"):
            arq_path = os.path.join(dir_mes, arq)
            parcial = obter_dados_por_classe(arq_path)
            for c in CLASSES:
                somar_linha(parcial[c], dados[c])
            contagem += 1

    if contagem == 0:
        messagebox.showwarning("Aviso", "Nenhuma planilha encontrada no mês informado.")
        return

    criar_relatorio(f"{mes}.xlsx", dados, os.path.join(BASE_DIR, ano), contagem)
    messagebox.showinfo("Sucesso", f"Relatório mensal de {mes} gerado com sucesso.")

def gerar_trimestral():
    ano = entry_ano.get()
    dir_ano = os.path.join(BASE_DIR, ano)
    for trimestre, meses in TRIMESTRES.items():
        dados = {c: inicializa_linha() for c in CLASSES}
        contagem = 0
        for mes in meses:
            dir_mes = os.path.join(dir_ano, mes)
            if not os.path.exists(dir_mes):
                continue
            for arq in os.listdir(dir_mes):
                if arq.endswith(".xlsx"):
                    arq_path = os.path.join(dir_mes, arq)
                    parcial = obter_dados_por_classe(arq_path)
                    for c in CLASSES:
                        somar_linha(parcial[c], dados[c])
                    contagem += 1
        if contagem > 0:
            criar_relatorio(f"{trimestre}.xlsx", dados, dir_ano, contagem)
    messagebox.showinfo("Sucesso", f"Relatórios trimestrais de {ano} gerados com sucesso.")

def gerar_anual():
    ano = entry_ano.get()
    dir_ano = os.path.join(BASE_DIR, ano)
    dados = {c: inicializa_linha() for c in CLASSES}
    contagem = 0
    for mes in MESES:
        dir_mes = os.path.join(dir_ano, mes)
        if not os.path.exists(dir_mes):
            continue
        for arq in os.listdir(dir_mes):
            if arq.endswith(".xlsx"):
                arq_path = os.path.join(dir_mes, arq)
                parcial = obter_dados_por_classe(arq_path)
                for c in CLASSES:
                    somar_linha(parcial[c], dados[c])
                contagem += 1
    if contagem > 0:
        criar_relatorio("Relatório_Anual.xlsx", dados, dir_ano, contagem)
        messagebox.showinfo("Sucesso", f"Relatório anual de {ano} gerado com sucesso.")

# Interface
janela = tk.Tk()
janela.title("Gerador de Relatórios da EBD")

frame = ttk.Frame(janela, padding=20)
frame.pack()

label_ano = ttk.Label(frame, text="Ano:")
label_ano.grid(row=0, column=0, sticky="w")
entry_ano = ttk.Combobox(frame, values=[str(a) for a in range(2020, 2031)])
entry_ano.grid(row=0, column=1)
entry_ano.set("2024")

label_mes = ttk.Label(frame, text="Mês:")
label_mes.grid(row=1, column=0, sticky="w")
combo_mes = ttk.Combobox(frame, values=MESES)
combo_mes.grid(row=1, column=1)
combo_mes.set("janeiro")

botao_mensal = ttk.Button(frame, text="Gerar Relatório Mensal", command=gerar_mensal)
botao_mensal.grid(row=2, column=0, columnspan=2, pady=5)

botao_trimestral = ttk.Button(frame, text="Gerar Trimestrais", command=gerar_trimestral)
botao_trimestral.grid(row=3, column=0, columnspan=2, pady=5)

botao_anual = ttk.Button(frame, text="Gerar Anual", command=gerar_anual)
botao_anual.grid(row=4, column=0, columnspan=2, pady=5)

janela.mainloop()
